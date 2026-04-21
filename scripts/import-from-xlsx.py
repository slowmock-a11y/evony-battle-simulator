#!/usr/bin/env python3
"""
Enrich data/generals.json and produce data/covenants.json from the
"Kopie von Master Generals Spreadsheet2.xlsx" workbook.

Usage:
    python scripts/import-from-xlsx.py [path/to/workbook.xlsx]

Requires: openpyxl
After running:  node scripts/regen-generals-data.js
"""

import json
import os
import sys
import warnings
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.stderr.write("openpyxl is required: pip install openpyxl\n")
    sys.exit(1)

warnings.filterwarnings("ignore")

REPO_ROOT = Path(__file__).resolve().parent.parent
GENERALS_JSON = REPO_ROOT / "data" / "generals.json"
COVENANTS_JSON = REPO_ROOT / "data" / "covenants.json"

DEFAULT_WORKBOOK = Path(
    os.environ.get("HOME", os.environ.get("USERPROFILE", ""))
) / "Downloads" / "Kopie von Master Generals Spreadsheet2.xlsx"

TROOP_TYPES = ("mounted", "ground", "ranged", "siege")


# ---------- xlsx parsing helpers ----------

def _num(v):
    """Coerce a cell value to int/float/None. Non-numeric (formulas etc.) → None."""
    if v is None or v == "" or v == "#REF!":
        return None
    if isinstance(v, bool):
        return int(v)
    if isinstance(v, (int, float)):
        return v
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _buff_block(ws, row, start_col):
    """Read a 12-cell block (4 troop types × ATK/DEF/HP) into a nested dict."""
    out = {}
    for i, t in enumerate(TROOP_TYPES):
        base = start_col + i * 3
        out[t] = {
            "atk": _num(ws.cell(row, base).value) or 0,
            "def": _num(ws.cell(row, base + 1).value) or 0,
            "hp":  _num(ws.cell(row, base + 2).value) or 0,
        }
    return out


def _norm_name(s):
    """Normalize a general's name for matching (strip leading '*', whitespace)."""
    if s is None:
        return ""
    return str(s).strip().lstrip("*").strip()


# ---------- sheet readers ----------

def read_sub_city(ws):
    """
    SUB-CITY sheet: data rows 4..end; name col 2; Total buffs cols 51-62;
    per-troop-type debuff overall cols 64-67 (Mounted/Ground/Ranged/Siege).
    """
    out = {}
    for r in range(4, ws.max_row + 1):
        name = _norm_name(ws.cell(r, 2).value)
        if not name:
            continue
        buffs = _buff_block(ws, r, 51)
        debuffs = {}
        for i, t in enumerate(TROOP_TYPES):
            debuffs[t] = _num(ws.cell(r, 64 + i).value) or 0
        out[name] = {
            "sub_city_buffs": buffs,
            "sub_city_debuffs": debuffs,
        }
    return out


def read_duty(ws):
    """
    DUTY sheet: data rows 5..end; name col 2.
    Overall defending: Rally Capacity col 3, March Size col 4.
    Overall attacking: Rally Capacity col 15, March Size col 16.
    Detailed Total blocks (Full Ascension + Full Specialties + Full Covenant/Skin):
      - Defending BUFFS:   cols 75-86
      - Defending DEBUFFS: cols 135-146
      - Attacking BUFFS:   cols 195-206
      - Attacking DEBUFFS: cols 255-266

    Returns (data_by_name, has_any_data). In the current spreadsheet the DUTY
    sheet is a template — only names are filled in — so we detect the empty
    state and skip merging if so.
    """
    out = {}
    has_any = False
    for r in range(5, ws.max_row + 1):
        name = _norm_name(ws.cell(r, 2).value)
        if not name:
            continue
        rally_def = _num(ws.cell(r, 3).value)
        march_def = _num(ws.cell(r, 4).value)
        rally_att = _num(ws.cell(r, 15).value)
        march_att = _num(ws.cell(r, 16).value)
        def_buffs   = _buff_block(ws, r, 75)
        def_debuffs = _buff_block(ws, r, 135)
        att_buffs   = _buff_block(ws, r, 195)
        att_debuffs = _buff_block(ws, r, 255)
        if any([rally_def, march_def, rally_att, march_att]) or \
           any(v for b in (def_buffs, def_debuffs, att_buffs, att_debuffs)
                 for t in TROOP_TYPES for v in b[t].values()):
            has_any = True
        out[name] = {
            "duty_rally_capacity_pct_defending": rally_def or 0,
            "duty_march_size_pct_defending":     march_def or 0,
            "duty_rally_capacity_pct_attacking": rally_att or 0,
            "duty_march_size_pct_attacking":     march_att or 0,
            "duty_defending_buffs":   def_buffs,
            "duty_defending_debuffs": def_debuffs,
            "duty_attacking_buffs":   att_buffs,
            "duty_attacking_debuffs": att_debuffs,
        }
    return out, has_any


def read_wall(ws):
    """
    WALL sheet: data rows 3..end; name col 2 (may start with '*');
    Total block at cols 51-62. First occurrence wins (sheet has duplicates).
    """
    out = {}
    for r in range(3, ws.max_row + 1):
        name = _norm_name(ws.cell(r, 2).value)
        if not name or name in out:
            continue
        out[name] = {"wall_buffs": _buff_block(ws, r, 51)}
    return out


def read_covenants(ws):
    """
    COVENANTS sheet: data rows 2..74 (73 rows visible, header excluded).
    Cols A-D: Main / Covenant I / II / III. Cols E-F: Gold Buff 1 / 2.
    """
    out = []
    for r in range(2, ws.max_row + 1):
        main = _norm_name(ws.cell(r, 1).value)
        if not main:
            continue
        members = [
            _norm_name(ws.cell(r, 2).value),
            _norm_name(ws.cell(r, 3).value),
            _norm_name(ws.cell(r, 4).value),
        ]
        members = [m for m in members if m]
        gold_buffs = []
        for c in (5, 6):
            v = ws.cell(r, c).value
            if v is not None and str(v).strip():
                gold_buffs.append(str(v).strip())
        out.append({
            "main": main,
            "members": members,
            "gold_buffs": gold_buffs,
        })
    return out


# ---------- merging ----------

def merge_sub_city(generals_by_name, sub_city_data):
    matched, unmatched = 0, []
    for name, fields in sub_city_data.items():
        entry = generals_by_name.get(name)
        if not entry:
            unmatched.append(name)
            continue
        entry["sub_city_buffs"] = fields["sub_city_buffs"]
        entry["sub_city_debuffs"] = fields["sub_city_debuffs"]
        matched += 1
    return matched, unmatched


def merge_duty(generals_by_name, duty_data):
    matched, unmatched = 0, []
    for name, fields in duty_data.items():
        entry = generals_by_name.get(name)
        if not entry:
            unmatched.append(name)
            continue
        entry.update({
            "duty_rally_capacity_pct_defending": fields["duty_rally_capacity_pct_defending"],
            "duty_march_size_pct_defending":     fields["duty_march_size_pct_defending"],
            "duty_rally_capacity_pct_attacking": fields["duty_rally_capacity_pct_attacking"],
            "duty_march_size_pct_attacking":     fields["duty_march_size_pct_attacking"],
            "duty_defending_buffs":   fields["duty_defending_buffs"],
            "duty_defending_debuffs": fields["duty_defending_debuffs"],
            "duty_attacking_buffs":   fields["duty_attacking_buffs"],
            "duty_attacking_debuffs": fields["duty_attacking_debuffs"],
        })
        matched += 1
    return matched, unmatched


def merge_wall(generals_list, generals_by_name, wall_data):
    """
    For generals already in JSON: attach wall_buffs.
    For wall-only names: create a minimal new entry with troop_type='WALL'.
    """
    attached = 0
    new_entries = 0
    for name, fields in wall_data.items():
        entry = generals_by_name.get(name)
        if entry:
            entry["wall_buffs"] = fields["wall_buffs"]
            attached += 1
        else:
            new = {
                "name": name,
                "troop_type": "WALL",
                "type": None,
                "role": "wall",
                "tavern": None,
                "main_skill_atk_pct": None,
                "main_skill_def_pct": None,
                "main_skill_hp_pct": None,
                "main_skill_march_size_pct": None,
                "wall_buffs": fields["wall_buffs"],
            }
            generals_list.append(new)
            generals_by_name[name] = new
            new_entries += 1
    return attached, new_entries


def link_covenants(generals_by_name, covenants):
    """
    Add `covenants` back-reference (list of main-general names) to each
    general appearing in a covenant. A general can belong to multiple covenants.
    """
    linked, unmatched = 0, set()
    for cov in covenants:
        for member in [cov["main"]] + cov["members"]:
            entry = generals_by_name.get(member)
            if entry:
                entry.setdefault("covenants", [])
                if cov["main"] not in entry["covenants"]:
                    entry["covenants"].append(cov["main"])
                linked += 1
            else:
                unmatched.add(member)
    return linked, sorted(unmatched)


# ---------- main ----------

def main():
    xlsx_path = Path(sys.argv[1]) if len(sys.argv) > 1 else DEFAULT_WORKBOOK
    if not xlsx_path.exists():
        sys.stderr.write(f"workbook not found: {xlsx_path}\n")
        sys.exit(1)

    print(f"Reading {xlsx_path}")
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)

    sub_city_data = read_sub_city(wb["SUB-CITY"])
    duty_data, duty_has_data = read_duty(wb["DUTY"])
    wall_data = read_wall(wb["WALL"])
    covenants = read_covenants(wb["COVENANTS"])

    print(f"  SUB-CITY: {len(sub_city_data)} generals")
    print(f"  DUTY:     {len(duty_data)} generals (has_data={duty_has_data})")
    print(f"  WALL:     {len(wall_data)} unique generals")
    print(f"  COVENANTS:{len(covenants)} entries")

    with GENERALS_JSON.open(encoding="utf-8") as f:
        generals = json.load(f)

    # Idempotency: drop previously-imported fields before re-merging, and
    # remove entries we created ourselves (troop_type == 'WALL' with role == 'wall').
    imported_keys = (
        "sub_city_buffs", "sub_city_debuffs",
        "duty_defending_buffs", "duty_defending_debuffs",
        "duty_attacking_buffs", "duty_attacking_debuffs",
        "duty_rally_capacity_pct_defending", "duty_march_size_pct_defending",
        "duty_rally_capacity_pct_attacking", "duty_march_size_pct_attacking",
        "wall_buffs", "covenant", "covenants",
    )
    generals[:] = [g for g in generals
                   if not (g.get("troop_type") == "WALL" and g.get("role") == "wall")]
    for g in generals:
        for k in imported_keys:
            g.pop(k, None)
    by_name = {g["name"]: g for g in generals}

    m1, u1 = merge_sub_city(by_name, sub_city_data)
    if duty_has_data:
        m2, u2 = merge_duty(by_name, duty_data)
    else:
        print("  [skip] DUTY sheet contains no numeric data — skipping merge.")
        m2, u2 = 0, []
    m3, new_w = merge_wall(generals, by_name, wall_data)
    m4, u4 = link_covenants(by_name, covenants)

    print(f"\nMerge report:")
    print(f"  SUB_CITY matched : {m1}/{len(sub_city_data)}  unmatched: {u1}")
    print(f"  DUTY matched     : {m2}/{len(duty_data)}  unmatched: {u2}")
    print(f"  WALL attached    : {m3}  new WALL entries: {new_w}")
    print(f"  Covenant links   : {m4}  unmatched members: {len(u4)}")
    if u4:
        print(f"    first few: {u4[:10]}")

    # Fail loud if the fill-in sheets didn't match — we want to catch name drift.
    if u1 or (duty_has_data and u2):
        sys.stderr.write("\nUnmatched SUB_CITY/DUTY names — aborting so you can fix aliases.\n")
        sys.exit(1)

    # Write generals.json with a trailing newline, stable 2-space indent (matches existing).
    with GENERALS_JSON.open("w", encoding="utf-8") as f:
        json.dump(generals, f, indent=2, ensure_ascii=False)
        f.write("\n")
    print(f"\nWrote {GENERALS_JSON}  ({len(generals)} entries)")

    with COVENANTS_JSON.open("w", encoding="utf-8") as f:
        json.dump(covenants, f, indent=2, ensure_ascii=False)
        f.write("\n")
    print(f"Wrote {COVENANTS_JSON} ({len(covenants)} entries)")

    print("\nNext: node scripts/regen-generals-data.js")


if __name__ == "__main__":
    main()
